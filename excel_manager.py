# Abbreviated as XLM (Excel Mangager)

# Nonlocal Modules
import openpyxl
from openpyxl import load_workbook
import os
import sys
import codecs
import hashlib
import csv


# Create an excel file
def create_file(wb_name):
    if '.xlsx' not in wb_name:
        wb_name += '.xlsx'
    path = os.getcwd()+'/'+wb_name
    wb = openpyxl.Workbook()
    wb.save(path)

# Get a local XLSX file and return in an object
def get_workbook(wb_name):
    return load_workbook(str(wb_name), data_only=True)

# Get a XLSX file's sheet
def get_worksheet(wb, ws_name):
    return wb[str(ws_name)]

# Add a sheet to an XLSX file
def add_worksheet(wb, ws_name):
    return wb.create_sheet(ws_name)

# Add a sheet to an XLSX file
def delete_worksheet(wb, ws_name):
    del wb[ws_name]

# Get an Excel file's sheet names
def sheet_names(wb):
    return wb.sheetnames

def rename_worksheet(wb, filename, original, new):
    ws = get_worksheet(wb, original)
    ws.title = new
    wb.save(filename)

def save_workbook(wb, filename):
    wb.save(filename)


# --------------------------------------------- #


# Returns number of rows, starting at 1
def num_rows(ws):
    return ws.max_row

# Returns number of columns, starting at 1
def num_columns(ws):
    return ws.max_column


# --------------------------------------------- #


# Returns all values of a single row in list format
def get_row(ws, row_num, col_start=1, col_end=0):
    if col_end == 0:
        col_end = num_columns(ws) + 1
    elif col_end < 0:
        col_end = num_columns(ws) + 1 + col_end
    row = []
    for i in range(col_start, col_end):
        row.append(get_cell(ws, i, row_num))
    return row

# Returns all values of a single column in list format
def get_column(ws, col_num, row_start=1, row_end=0):
    if row_end == 0:
        row_end = num_rows(ws) + 1
    elif row_end < 0:
        row_end = num_rows(ws) + 1 + row_end
    column = []
    for i in range(row_start, row_end):
        column.append(get_cell(ws, col_num, i))
    return column

# Returns the value of a single cell
def get_cell(ws, col_num, row_num):
    cell = ws.cell(column=col_num, row=row_num).value
    if isinstance(cell, str):
        cell = cell.strip()
        if cell == '':
            cell = None
    return cell


# --------------------------------------------- #


# Edits the value of a single cell
def edit_cell(ws, col_num, row_num, new_value):
    ws.cell(column=col_num, row=row_num).value = new_value


# --------------------------------------------- #


def string_row_pos(ws, string, col_num, row_start=1, row_end=0):
    if row_end == 0:
        row_end = num_rows(ws) + 1
    elif row_end < 0:
        row_end = num_rows(ws) + 1 + row_end
    for i in range(row_start, row_end):
        if get_cell(ws, col_num, i) == string:
            return i
    return -1

def string_col_pos(ws, string, row_num, col_start=1, col_end=0):
    if col_end == 0:
        col_end = num_rows(ws) + 1
    elif col_end < 0:
        col_end = num_rows(ws) + 1 + col_end
    for i in range(col_start, col_end):
        if get_cell(ws, i, row_num) == string:
            return i
    return -1


# --------------------------------------------- #


# Returns list of dictionaries, w/ each row being represented by a dictionary
# If col_headers is not passed as param, will assume all headers are on row 1
def rows_dict(ws, col_headers=[], row_start=2, row_end=0, col_start=1, col_end=0):
    if row_end == 0:
        row_end = num_rows(ws) + 1
    elif row_end < 0:
        row_end = num_rows(ws) + 1 + row_end
    if col_end == 0:
        col_end = num_columns(ws) + 1
    elif col_end < 0:
        col_end = num_columns(ws) + 1 + col_end
    if len(col_headers) == 0:
        col_headers = get_row(ws,1)
    elif len(col_headers) != num_columns(ws) - (col_start + 1):
        print('XLM ERROR: Passed parameter col_headers has ' + \
        str(len(col_headers)) + ' value(s), should have ' + str(num_columns) + '.')
        sys.exit()
    dict = []
    for i in range(row_start, row_end):
        row = {}
        for j in range(col_start, col_end):
            row[str(col_headers[j-1])] = get_cell(ws, j, i)
        dict.append(row)
    return dict

# Returns list of lists, w/ each row being represented by each inner list
# Does not store column headers at all (Should be done in separate list)
def rows_list(ws, row_start=2, row_end=0, col_start=1, col_end=0):
    if row_end == 0:
        row_end = num_rows(ws) + 1
    elif row_end < 0:
        row_end = num_rows(ws) + 1 + row_end
    if col_end == 0:
        col_end = num_columns(ws) + 1
    elif col_end < 0:
        col_end = num_columns(ws) + 1 + col_end
    list = []
    for i in range(row_start, row_end):
        row = []
        for j in range(col_start, col_end):
            row.append(get_cell(ws, j, i))
        list.append(row)
    return list


# --------------------------------------------- #


# If selected row is empty
def row_empty(ws, row_num, col_start=1, col_end=0):
    if col_end == 0:
        col_end = num_rows(ws) + 1
    elif col_end < 0:
        col_end = num_rows(ws) + 1 + col_end
    for i in range(col_start, col_end):
        if get_cell(ws, i, row_num) != None:
            return False
    return True



# If selected column is empty
def col_empty(ws, col_num, row_start=1, row_end=0):
    if row_end == 0:
        row_end = num_rows(ws) + 1
    elif row_end < 0:
        row_end = num_rows(ws) + 1 + row_end
    for i in range(row_start, row_end):
        if get_cell(ws, col_num, i) != None:
            return False
    return True


# --------------------------------------------- #


# Remove any whitespace from the edges of strings,
# as well as delete any empty rows.
def prime_workbook(wb_name):
    wb = get_workbook(wb_name)
    for sheet_name in sheet_names(wb):
        sheet = get_worksheet(wb, sheet_name)
        for i in range(1,num_rows(sheet)+1):
            if row_empty(sheet, i):
                sheet.delete_rows(i,1)
            for j in range(1,num_columns(sheet)+1):
                cell = get_cell(sheet,j,i)
                if isinstance(cell, str):
                    edit_cell(sheet, j, i, cell.strip())
    wb.save(wb_name)


# Remove all whitespaces on edges of all string cell values
def strip_all(wb_name):
    wb = get_workbook(wb_name)
    for sheet_name in sheet_names(wb):
        sheet = get_worksheet(wb, sheet_name)
        for i in range(1,num_rows(sheet)+1):
            for j in range(1,num_columns(sheet)+1):
                cell = get_cell(sheet,j,i)
                if isinstance(cell, str):
                    edit_cell(sheet, j, i, cell.strip())
    wb.save(wb_name)

# Delete any rows that are empty
def delete_empty(wb_name):
    wb = get_workbook(wb_name)
    for sheet_name in sheet_names(wb):
        sheet = get_worksheet(wb, sheet_name)
        for i in range(1,num_rows(sheet)+1):
            if row_empty(sheet, i):
                sheet.delete_rows(i,1)
    wb.save(wb_name)

# --------------------------------------------- #

# Create a hash with respect to an excel file's string contents
def get_hash(excelfile):
    wb = load_workbook(str(excelfile), read_only=True)
    content = ''
    for sheet in sheet_names(wb):
        ws = get_worksheet(wb, sheet)
        for i in range(1,num_rows(ws)+1):
            for j in range(1,num_columns(ws)+1):
                content += str(get_cell(ws,j,i)) + "~/"
            content += "~~/"
        content += "~~~/"
    hash_object = hashlib.sha1(content.encode())
    hex_dig = hash_object.hexdigest()
    return str(hex_dig)

def convert_csv(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    with open(path) as f:
        reader = csv.reader(f, delimiter=':')
        for row in reader:
            ws.append(row)
    return wb, ws
